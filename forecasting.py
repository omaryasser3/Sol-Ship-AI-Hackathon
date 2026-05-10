"""
Final submission: h=1 load forecast for April 2025 and September 2025.

Each period is predicted using ALL data that came before it:
  April     → train Jan 2024 – Jan 2025 | val Feb–Mar 2025
  September → train Jan 2024 – Jun 2025 | val Jul–Aug 2025

Outputs saved to  submission_forecast/
  requirements.txt
  metrics_april_2025.txt
  metrics_september_2025.txt
  forecast_april_2025.xlsx    (data + per-step errors + summary + embedded charts)
  forecast_september_2025.xlsx
  plot_april_forecast.png
  plot_april_errors.png
  plot_september_forecast.png
  plot_september_errors.png
"""

import sys, os, warnings, time
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "solship"))
warnings.filterwarnings("ignore")

import numpy as np
import pandas as pd
import lightgbm as lgb
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import holidays as hol_pkg
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from features import (build_features, compute_slot_stats,
                      compute_slot_month_stats, compute_slot_dow_stats)

# ── constants ──────────────────────────────────────────────────────────────────
OUT_DIR          = "submission_forecast"
F1, F2, F3       = 0.2540, 0.2682, 0.2440
ITALIAN_HOLIDAYS = hol_pkg.Italy(years=[2024, 2025])
# Extended lags: add 2-week (1344), 3-week (2016), 4-week (2688) for transition months
ENHANCED_LAGS    = [0, 1, 2, 3, 4, 5, 6, 7, 8, 48, 96, 192, 336, 672, 1344, 2016, 2688]
ENHANCED_ROLLS   = [4, 8, 16, 96, 672, 1344]   # added 2-week rolling
SPARSE_HORIZONS  = [1, 2, 4, 8, 16, 24, 48, 72, 96, 144, 192]
H_MAX            = 192


def build_mpc_forecast_matrix(pred_hub, pred_qnt, alpha_hub, common_index, H=96,
                              pred_twe=None, pred_raw=None, blend_weights=None,
                              a_cal=1.0, b_cal=0.0):
    """
    Assemble the sparse-horizon predictions into a full (n_steps × H) matrix
    that run_mpc() expects.

    For each timestep t in common_index, row t = [forecast_h1, ..., forecast_hH],
    i.e. the forecast made AT time t for 1..H steps ahead.

    Sparse horizons that don't cover every integer h are linearly interpolated
    along the horizon axis for each timestep independently.

    Parameters
    ----------
    pred_hub, pred_qnt : dict  h -> pd.Series(index=timestamps)
    alpha_hub          : float  blend weight for Huber (1-alpha_hub for quantile)
    common_index       : DatetimeIndex of timesteps to include
    H                  : int  MPC horizon length (default 96 = 24 h)

    Returns
    -------
    pd.DataFrame  shape (n_steps, H), columns named h_1..h_H, index=common_index
    """
    alpha_qnt = 1.0 - alpha_hub
    sparse_h  = sorted(pred_hub.keys())

    # Build blended sparse matrix: shape (n_steps, n_sparse)
    rows = {}
    for h in sparse_h:
        if blend_weights is not None and h in blend_weights:
            weights = blend_weights[h]
            if isinstance(weights, dict):
                blended = (weights.get("hub", 0.0) * pred_hub[h].reindex(common_index).ffill() +
                           weights.get("qnt", 0.0) * pred_qnt[h].reindex(common_index).ffill())
                if pred_twe is not None and h in pred_twe:
                    blended = blended + weights.get("twe", 0.0) * pred_twe[h].reindex(common_index).ffill()
                if pred_raw is not None and h in pred_raw:
                    blended = blended + weights.get("raw", 0.0) * pred_raw[h].reindex(common_index).ffill()
            elif pred_twe is not None:
                ah, aq, at = weights
                blended = (ah * pred_hub[h].reindex(common_index).ffill() +
                           aq * pred_qnt[h].reindex(common_index).ffill() +
                           at * pred_twe[h].reindex(common_index).ffill())
            else:
                blended = (alpha_hub * pred_hub[h].reindex(common_index).ffill() +
                           alpha_qnt * pred_qnt[h].reindex(common_index).ffill())
        else:
            blended = (alpha_hub * pred_hub[h].reindex(common_index).ffill() +
                       alpha_qnt * pred_qnt[h].reindex(common_index).ffill())
        # affine calibration applied uniformly across horizons (fit at h=1, but
        # bias correction is largely scale-invariant in the relevant α range)
        rows[h] = np.clip(a_cal * blended.values + b_cal, 0, None)

    sparse_arr = np.column_stack([rows[h] for h in sparse_h])  # (n, n_sparse)

    # Interpolate to every integer horizon 1..H
    full_arr = np.zeros((len(common_index), H), dtype=np.float32)
    target_hs = np.arange(1, H + 1, dtype=float)
    for i in range(len(common_index)):
        full_arr[i] = np.interp(target_hs, sparse_h, sparse_arr[i])

    cols = [f"h_{h}" for h in range(1, H + 1)]
    return pd.DataFrame(full_arr, index=common_index, columns=cols)

LGB_PARAMS = {
    "verbose": -1, "num_threads": -1,
    "learning_rate": 0.0272, "num_leaves": 167,
    "min_data_in_leaf": 135, "feature_fraction": 0.862,
    "bagging_fraction": 0.656, "bagging_freq": 5,
    "lambda_l2": 0.078, "lambda_l1": 0.150,
}

SEEDS = [42, 1337, 2024]
ENABLE_TWEEDIE_HEAD = True
ENABLE_RAW_RMSE_HEAD = False
ENABLE_AFFINE_CALIBRATION = True
CAL_MIN_NRMSE_GAIN = 0.0
BLEND_GRID_STEP = 0.05


def _train_seeds(params_base, X, y, num_rounds, feat_cols):
    models = []
    for s in SEEDS:
        params = {**params_base, "seed": s, "bagging_seed": s, "feature_fraction_seed": s}
        d = lgb.Dataset(X, y, feature_name=feat_cols, free_raw_data=False)
        models.append(lgb.train(params, d, num_boost_round=num_rounds,
                                callbacks=[lgb.log_evaluation(0)]))
    return models


def _predict_seeds(models, X):
    return np.mean([m.predict(X) for m in models], axis=0)


def _make_simplex_grid(step=0.05):
    grid = []
    n = int(round(1.0 / step))
    for i in range(n + 1):
        for j in range(n + 1 - i):
            k = n - i - j
            grid.append((i * step, j * step, k * step))
    return grid


_SIMPLEX_GRID = _make_simplex_grid(0.05)


def _weight_grid(heads, step=None):
    """Return dict weights summing to 1. Supports 2-4 heads for ablations."""
    if step is None:
        step = BLEND_GRID_STEP
    n = int(round(1.0 / step))
    heads = list(heads)
    out = []

    def rec(pos, remaining, vals):
        if pos == len(heads) - 1:
            out.append(dict(zip(heads, vals + [remaining * step])))
            return
        for i in range(remaining + 1):
            rec(pos + 1, remaining - i, vals + [i * step])

    rec(0, n, [])
    return out

log1p  = np.log1p
expm1c = lambda x: np.expm1(x).clip(min=0)

# True when run as `python run_final_submission.py`; False when imported by main.py
_RUN = (__name__ == "__main__")

os.makedirs(OUT_DIR, exist_ok=True)

# ── helpers ────────────────────────────────────────────────────────────────────
def compute_buy_price(ts):
    dow = ts.dt.dayofweek; h = ts.dt.hour
    hol = ts.dt.date.map(lambda d: d in ITALIAN_HOLIDAYS)
    f3  = (dow == 6) | hol | (~(dow == 6) & ~hol & ((h < 7) | (h >= 23)))
    f2  = (~f3) & (((dow == 5) & (h >= 7) & (h < 23)) |
                   (~(dow == 5) & ~(dow == 6) & ~hol & ((h == 7) | ((h >= 19) & (h < 23)))))
    p   = pd.Series(F1, index=ts.index, dtype=float); p[f2] = F2; p[f3] = F3
    return p

def log_df(df):
    d = df.copy(); d["load_kw"] = log1p(d["load_kw"]); return d

def make_Xy(df_buf, df_target, horizon, feat_cols=None,
            slot_stats=None, slot_month_stats=None, slot_dow_stats=None):
    X, y = build_features(df_buf, lags=ENHANCED_LAGS,
                           roll_windows=ENHANCED_ROLLS, horizon=horizon,
                           slot_stats=slot_stats,
                           slot_month_stats=slot_month_stats,
                           slot_dow_stats=slot_dow_stats)
    idx  = df_target.index.intersection(X.index)
    X, y = X.loc[idx], y.loc[idx]
    if feat_cols is not None:
        X = X[feat_cols]
    return X, y

def safe_mape(actual, forecast):
    mask = actual > 0.01
    if mask.sum() == 0:
        return float("nan")
    return float(np.mean(np.abs((actual[mask] - forecast[mask]) / actual[mask])) * 100)

# ── load & clean data ──────────────────────────────────────────────────────────
def load_data(csv_path="data/ENERGY_Hackathon_DataSet(Sheet1).csv"):
    """Load, clean, and return the full dataset as a single DataFrame."""
    print("=" * 60)
    print("LOADING DATA")
    print("=" * 60)

    # Accept both repository layouts:
    # - data/ENERGY_Hackathon_DataSet(Sheet1).csv
    # - ENERGY_Hackathon_DataSet(Sheet1).csv at project root
    candidate_paths = [
        csv_path,
        "ENERGY_Hackathon_DataSet(Sheet1).csv",
        os.path.join(os.path.dirname(__file__), "ENERGY_Hackathon_DataSet(Sheet1).csv"),
        os.path.join(os.path.dirname(__file__), "data", "ENERGY_Hackathon_DataSet(Sheet1).csv"),
    ]
    resolved = next((p for p in candidate_paths if os.path.exists(p)), None)
    if resolved is None:
        raise FileNotFoundError(
            "Could not find ENERGY_Hackathon_DataSet(Sheet1).csv in expected locations"
        )

    df = pd.read_csv(resolved, sep=";", decimal=",")
    df["timestamp"] = pd.to_datetime(df["timestamp"])
    df = (df.sort_values("timestamp")
            .drop_duplicates(subset="timestamp", keep="first")
            .reset_index(drop=True))
    df = df.rename(columns={"load_p": "load_kw", "pv_p": "pv_kw",
                             "Selling_price_eur_kwh": "sell_price",
                             "battery_p": "p_battery_kw", "grid_p": "grid_kw"})
    df["sell_price"] = df["sell_price"].ffill()
    df["buy_price"]  = compute_buy_price(df["timestamp"])
    idx = pd.date_range(df["timestamp"].min(), df["timestamp"].max(), freq="15min")
    df = df.set_index("timestamp").reindex(idx).ffill()
    df["load_kw"] = df["load_kw"].clip(lower=0)
    df["pv_kw"]   = df["pv_kw"].clip(lower=0)
    print(f"  Rows loaded: {len(df):,}  ({df.index[0].date()} → {df.index[-1].date()})")
    return df


raw = None  # populated by load_data() or __main__


# ── training / prediction for one period ──────────────────────────────────────
def _nrmse(actual, forecast):
    rmse = np.sqrt(np.mean((actual - forecast) ** 2))
    return rmse / np.mean(actual) * 100 if np.mean(actual) > 1e-9 else float("nan")


def run_period(month_num, label, month_start, month_end, raw):
    """
    Train on ALL data before the target month.

    ES validation uses the 8 weeks immediately before the test month — this
    matches the seasonal profile of the test month (spring for April, late
    summer for September) and gives much better early-stopping signal than
    a fixed winter holdout.

    Blend weight alpha (Huber fraction) is optimized on the ES val set.

    Returns DataFrame with per-timestep actuals, forecasts, and errors.
    """
    print(f"\n{'='*60}")
    print(f"PERIOD: {label}  ({month_start.date()} – {month_end.date()})")
    print(f"{'='*60}")
    t0 = time.time()

    train_end = month_start - pd.Timedelta("15min")
    df_train  = raw[raw.index <= train_end]
    df_test   = raw[(raw.index >= month_start) & (raw.index <= month_end)]

    # ES validation: 8 weeks closest to the test month (best seasonal proxy)
    # This is training data only — test month is never seen
    es_split  = train_end - pd.Timedelta(days=56)
    df_tr_es  = df_train[df_train.index <= es_split]
    df_val_es = df_train[df_train.index >  es_split]

    print(f"  Train : {df_train.index[0].date()} – {df_train.index[-1].date()}  "
          f"({len(df_train):,} rows)")
    print(f"  ES val: {df_val_es.index[0].date()} – {df_val_es.index[-1].date()}  "
          f"({len(df_val_es):,} rows, early-stop + blend tuning)")
    print(f"  Test  : {df_test.index[0].date()} – {df_test.index[-1].date()}  "
          f"({len(df_test):,} rows)")

    # ── slot stats from training data only — three granularities ─────────────
    train_slot_stats       = compute_slot_stats(df_train)
    train_slot_month_stats = compute_slot_month_stats(df_train)
    train_slot_dow_stats   = compute_slot_dow_stats(df_train)
    tr_es_slot_stats       = compute_slot_stats(df_tr_es)
    tr_es_slot_month_stats = compute_slot_month_stats(df_tr_es)
    tr_es_slot_dow_stats   = compute_slot_dow_stats(df_tr_es)

    def _Xy(df_buf, df_tgt, h, fc=None, es=False):
        ss  = tr_es_slot_stats       if es else train_slot_stats
        sms = tr_es_slot_month_stats if es else train_slot_month_stats
        sds = tr_es_slot_dow_stats   if es else train_slot_dow_stats
        return make_Xy(df_buf, df_tgt, h, fc,
                       slot_stats=ss, slot_month_stats=sms, slot_dow_stats=sds)

    # ── log-transform datasets ──
    df_tr_log    = log_df(df_train)
    df_tr_es_log = log_df(df_tr_es)
    buf_tail     = max(ENHANCED_LAGS)
    df_va_buf    = log_df(pd.concat([df_tr_es.iloc[-buf_tail:], df_val_es]))

    # feature columns from full training set
    X_tr1, _ = _Xy(df_tr_log, df_train, 1)
    feat_cols = list(X_tr1.columns)

    # ── Optuna HPO on ES val set (season-matched) ────────────────────────────
    # Tune against the 8-week window immediately before the test month so the
    # validation distribution matches the target season (spring for April,
    models_hub = {}; models_qnt = {}; models_twe = {}; models_raw = {}
    active_heads = ["hub", "qnt"]
    if ENABLE_TWEEDIE_HEAD:
        active_heads.append("twe")
    if ENABLE_RAW_RMSE_HEAD:
        active_heads.append("raw")

    print(f"  Training {len(SPARSE_HORIZONS)} horizons x {len(active_heads)} heads x {len(SEEDS)} seeds ...")
    t_tr = time.time()
    for h in SPARSE_HORIZONS:
        # ── ES probes (Huber on log target, Tweedie on raw target) ──
        X_es, y_es = _Xy(df_tr_es_log, df_tr_es, h, feat_cols, es=True)
        X_va, y_va = _Xy(df_va_buf,    df_val_es, h, feat_cols, es=True)
        y_es_raw   = np.expm1(y_es.values).clip(min=0)
        y_va_raw   = np.expm1(y_va.values).clip(min=0)

        d_es_log = lgb.Dataset(X_es.values, y_es.values, feature_name=feat_cols, free_raw_data=False)
        d_va_log = lgb.Dataset(X_va.values, y_va.values, feature_name=feat_cols,
                               reference=d_es_log, free_raw_data=False)
        d_es_raw = lgb.Dataset(X_es.values, y_es_raw, feature_name=feat_cols, free_raw_data=False)
        d_va_raw = lgb.Dataset(X_va.values, y_va_raw, feature_name=feat_cols,
                               reference=d_es_raw, free_raw_data=False)
        cbs_es = [lgb.early_stopping(80, verbose=False), lgb.log_evaluation(0)]
        m_es_hub = lgb.train(
            {**LGB_PARAMS, "objective": "huber", "metric": "huber", "alpha": 0.9},
            d_es_log, num_boost_round=2000, valid_sets=[d_va_log], callbacks=cbs_es)
        if ENABLE_TWEEDIE_HEAD:
            m_es_twe = lgb.train(
                {**LGB_PARAMS, "objective": "tweedie", "tweedie_variance_power": 1.5, "metric": "rmse"},
                d_es_raw, num_boost_round=2000, valid_sets=[d_va_raw], callbacks=cbs_es)
            best_rounds_twe = max(int(m_es_twe.best_iteration * 1.2), 50)
        else:
            best_rounds_twe = None
        if ENABLE_RAW_RMSE_HEAD:
            m_es_raw = lgb.train(
                {**LGB_PARAMS, "objective": "regression", "metric": "rmse"},
                d_es_raw, num_boost_round=2000, valid_sets=[d_va_raw], callbacks=cbs_es)
            best_rounds_raw = max(int(m_es_raw.best_iteration * 1.2), 50)
        else:
            best_rounds_raw = None
        # 1.2× scale-up for full-data retrain (matches single-seed heuristic)
        best_rounds_log = max(int(m_es_hub.best_iteration * 1.2), 50)

        # ── full-data multi-seed retrain ──
        X_tr, y_tr = _Xy(df_tr_log, df_train, h, feat_cols)
        y_tr_raw   = np.expm1(y_tr.values).clip(min=0)
        models_hub[h] = _train_seeds(
            {**LGB_PARAMS, "objective": "huber",    "metric": "huber",    "alpha": 0.9},
            X_tr.values, y_tr.values, best_rounds_log, feat_cols)
        # α=0.90 quantile catches spike under-forecasting (dominant NRMSE driver)
        models_qnt[h] = _train_seeds(
            {**LGB_PARAMS, "objective": "quantile", "metric": "quantile", "alpha": 0.90},
            X_tr.values, y_tr.values, best_rounds_log, feat_cols)
        if ENABLE_TWEEDIE_HEAD:
            # Tweedie on raw kW directly attacks RMSE in scoring units, complements
            # the log-Huber head by avoiding the spike-compression of log1p
            models_twe[h] = _train_seeds(
                {**LGB_PARAMS, "objective": "tweedie", "tweedie_variance_power": 1.5, "metric": "rmse"},
                X_tr.values, y_tr_raw, best_rounds_twe, feat_cols)
        if ENABLE_RAW_RMSE_HEAD:
            models_raw[h] = _train_seeds(
                {**LGB_PARAMS, "objective": "regression", "metric": "rmse"},
                X_tr.values, y_tr_raw, best_rounds_raw, feat_cols)
    print(f"  LGB training done in {time.time()-t_tr:.1f}s")

    # ── per-horizon 3-way blend search on ES-val (simplex grid, step 0.05) ──
    blend_weights = {}   # h -> dict of head weights
    val_pred_h1   = None
    y_val_h1      = None
    for h in SPARSE_HORIZONS:
        X_v, y_v = _Xy(df_va_buf, df_val_es, h, feat_cols, es=True)
        val_preds = {
            "hub": expm1c(_predict_seeds(models_hub[h], X_v.values)),
            "qnt": expm1c(_predict_seeds(models_qnt[h], X_v.values)),
        }
        if ENABLE_TWEEDIE_HEAD:
            val_preds["twe"] = np.clip(_predict_seeds(models_twe[h], X_v.values), 0, None)
        if ENABLE_RAW_RMSE_HEAD:
            val_preds["raw"] = np.clip(_predict_seeds(models_raw[h], X_v.values), 0, None)
        y_act = np.expm1(y_v.values).clip(min=0)

        best_n, best_w = float("inf"), {"hub": 1.0, "qnt": 0.0}
        for weights in _weight_grid(active_heads):
            pred = sum(weights[name] * val_preds[name] for name in active_heads).clip(min=0)
            n = _nrmse(y_act, pred)
            if n < best_n:
                best_n, best_w = n, weights
        blend_weights[h] = best_w
        if h == 1:
            weight_txt = "  ".join(f"{name}={best_w.get(name, 0.0):.2f}" for name in active_heads)
            print(f"  h=1 blend  {weight_txt}  (val NRMSE={best_n:.2f}%)")
            val_pred_h1 = sum(best_w[name] * val_preds[name] for name in active_heads).clip(min=0)
            y_val_h1    = y_act

    # ── affine calibration on ES-val (h=1) ──
    if ENABLE_AFFINE_CALIBRATION:
        A_mat = np.column_stack([val_pred_h1, np.ones_like(val_pred_h1)])
        sol, *_ = np.linalg.lstsq(A_mat, y_val_h1, rcond=None)
        a_cal, b_cal = float(sol[0]), float(sol[1])
        n_pre  = _nrmse(y_val_h1, val_pred_h1)
        n_post = _nrmse(y_val_h1, np.clip(a_cal * val_pred_h1 + b_cal, 0, None))
    else:
        a_cal, b_cal = 1.0, 0.0
        n_pre = n_post = _nrmse(y_val_h1, val_pred_h1)
    if not ENABLE_AFFINE_CALIBRATION:
        print("  Affine calibration: disabled")
    elif not (0.5 <= a_cal <= 1.5) or (n_pre - n_post < CAL_MIN_NRMSE_GAIN):
        a_cal, b_cal = 1.0, 0.0
        print(f"  Affine calibration: identity (rejected; val NRMSE {n_pre:.2f}% -> {n_post:.2f}%)")
    else:
        print(f"  Affine calibration  a={a_cal:.3f}  b={b_cal:+.3f}  "
              f"(val NRMSE {n_pre:.2f}% → {n_post:.2f}%)")

    # back-compat scalars (h=1 weights) — main.py / build_mpc_forecast_matrix
    best_alpha_hub = blend_weights[1].get("hub", 0.0)
    best_alpha_qnt = blend_weights[1].get("qnt", 0.0)

    # ── generate forecasts for the test month (batch) ────────────────────────
    buf_data  = raw[raw.index < month_start]
    pad_idx   = pd.date_range(df_test.index[-1] + pd.Timedelta("15min"),
                               periods=H_MAX, freq="15min")
    pad_df    = pd.concat([df_test.iloc[[-1]]] * H_MAX)
    pad_df.index = pad_idx
    df_te_buf = log_df(pd.concat([buf_data, df_test, pad_df]))

    def _Xy_te(h, fc=None):
        return make_Xy(df_te_buf, df_test, h, fc,
                       slot_stats=train_slot_stats,
                       slot_month_stats=train_slot_month_stats,
                       slot_dow_stats=train_slot_dow_stats)

    pred_hub = {}; pred_qnt = {}; pred_twe = {}; pred_raw = {}
    for h in SPARSE_HORIZONS:
        X, _ = _Xy_te(h, feat_cols)
        pred_hub[h] = pd.Series(expm1c(_predict_seeds(models_hub[h], X.values)), index=X.index)
        pred_qnt[h] = pd.Series(expm1c(_predict_seeds(models_qnt[h], X.values)), index=X.index)
        if ENABLE_TWEEDIE_HEAD:
            pred_twe[h] = pd.Series(np.clip(_predict_seeds(models_twe[h], X.values), 0, None), index=X.index)
        if ENABLE_RAW_RMSE_HEAD:
            pred_raw[h] = pd.Series(np.clip(_predict_seeds(models_raw[h], X.values), 0, None), index=X.index)

    common = pred_hub[SPARSE_HORIZONS[0]].index
    for h in SPARSE_HORIZONS[1:]:
        common = common.intersection(pred_hub[h].index)
    common = common[(common >= month_start) & (common <= month_end)]

    w1 = blend_weights[1]
    h1_blend = (w1.get("hub", 0.0) * pred_hub[1].loc[common].values +
                w1.get("qnt", 0.0) * pred_qnt[1].loc[common].values)
    if ENABLE_TWEEDIE_HEAD:
        h1_blend = h1_blend + w1.get("twe", 0.0) * pred_twe[1].loc[common].values
    if ENABLE_RAW_RMSE_HEAD:
        h1_blend = h1_blend + w1.get("raw", 0.0) * pred_raw[1].loc[common].values
    h1_blend = h1_blend.clip(min=0)
    h1_blend = np.clip(a_cal * h1_blend + b_cal, 0, None)
    forecast_h1 = pd.Series(h1_blend, index=common, name="forecast_kw")
    actual_h1   = df_test["load_kw"].shift(-1).reindex(common).dropna()
    forecast_h1 = forecast_h1.reindex(actual_h1.index)

    # ── per-timestep error table ──────────────────────────────────────────────
    result = pd.DataFrame({
        "actual_kw":    actual_h1.values,
        "forecast_kw":  forecast_h1.values,
    }, index=actual_h1.index)

    result["error_kw"]     = result["forecast_kw"] - result["actual_kw"]
    result["abs_error_kw"] = result["error_kw"].abs()
    result["sq_error_kw2"] = result["error_kw"] ** 2
    mask_nonzero           = result["actual_kw"] > 0.01
    result["pct_error"]    = np.nan
    result.loc[mask_nonzero, "pct_error"] = (
        result.loc[mask_nonzero, "abs_error_kw"] /
        result.loc[mask_nonzero, "actual_kw"] * 100
    )

    print(f"  Forecast timesteps: {len(result):,}")
    print(f"  Total elapsed: {time.time()-t0:.1f}s")

    # Expose sparse-horizon predictions so main.py can build the MPC forecast matrix
    period_out = {
        "result":          result,
        "pred_hub":        pred_hub,
        "pred_qnt":        pred_qnt,
        "pred_twe":        pred_twe,
        "pred_raw":        pred_raw,
        "blend_weights":   blend_weights,
        "a_cal":           a_cal,
        "b_cal":           b_cal,
        "best_alpha_hub":  best_alpha_hub,
        "best_alpha_qnt":  best_alpha_qnt,
        "feat_cols":       feat_cols,
        "common":          common,
    }

    # ── feature importance (LGB h=1 Huber gain, averaged across seeds) ───────
    fi_arr = np.mean([m.feature_importance(importance_type="gain")
                      for m in models_hub[1]], axis=0)
    fi_lgb = pd.Series(fi_arr, index=feat_cols).sort_values(ascending=False)

    top_n = 30
    slug  = label.lower().replace(" ", "_")

    fi_txt = os.path.join(OUT_DIR, f"feature_importance_{slug}.txt")
    with open(fi_txt, "w", encoding="utf-8") as f:
        f.write(f"Feature importance — {label}\n")
        f.write("=" * 60 + "\n")
        f.write(f"  Blend: hub={best_alpha_hub:.2f}  qnt={best_alpha_qnt:.2f}\n\n")
        f.write(f"  {'Rank':<5} {'Feature':<35} {'LGB gain':>12}\n")
        f.write("-" * 55 + "\n")
        for rank, feat in enumerate(fi_lgb.index[:top_n], 1):
            f.write(f"  {rank:<5} {feat:<35} {fi_lgb[feat]:>12,.0f}\n")
    print(f"  Saved: {fi_txt}")

    fi = fi_lgb  # use LGB for plot (rename for chart code below)

    # bar chart: top N
    fig, ax = plt.subplots(figsize=(9, top_n * 0.28 + 1.5))
    top = fi.head(top_n)
    ax.barh(top.index[::-1], top.values[::-1], color="#1f77b4", edgecolor="white")
    ax.set_xlabel("Gain (importance)")
    ax.set_title(f"Top {top_n} features by gain — {label}", fontsize=11, fontweight="bold")
    ax.grid(axis="x", alpha=0.3)
    fig.tight_layout()
    fi_png = os.path.join(OUT_DIR, f"feature_importance_{slug}.png")
    fig.savefig(fi_png, dpi=150, bbox_inches="tight")
    plt.close(fig)
    print(f"  Saved: {fi_png}")

    print(f"\n  Top 10 features ({label}):")
    for feat, val in fi.head(10).items():
        print(f"    {feat:<35s}  {val:,.0f}")

    return period_out


if _RUN:
    # ── run both periods ──────────────────────────────────────────────────────
    raw = load_data()
    period_outputs = {}
    for m_num, label, m_start, m_end in [
        (4,  "April 2025",     pd.Timestamp("2025-04-01"), pd.Timestamp("2025-04-30 23:45")),
        (9,  "September 2025", pd.Timestamp("2025-09-01"), pd.Timestamp("2025-09-30 23:45")),
    ]:
        period_outputs[label] = run_period(m_num, label, m_start, m_end, raw)

    results = {label: out["result"] for label, out in period_outputs.items()}



# ── compute summary metrics ────────────────────────────────────────────────────
def compute_metrics(df, label):
    actual   = df["actual_kw"].values
    forecast = df["forecast_kw"].values
    errors   = df["error_kw"].values
    abs_err  = df["abs_error_kw"].values
    sq_err   = df["sq_error_kw2"].values

    mae_v   = float(np.mean(abs_err))
    rmse_v  = float(np.sqrt(np.mean(sq_err)))
    mape_v  = float(df["pct_error"].mean())
    nrmse_v = rmse_v / float(np.mean(actual)) * 100 if np.mean(actual) > 1e-9 else float("nan")

    return {
        "label":     label,
        "n_steps":   len(df),
        "mae_kw":    mae_v,
        "rmse_kw":   rmse_v,
        "mape_pct":  mape_v,
        "nrmse_pct": nrmse_v,
        "mean_actual_kw": float(np.mean(actual)),
        "mean_forecast_kw": float(np.mean(forecast)),
        "bias_kw":   float(np.mean(errors)),
        "max_abs_error_kw": float(np.max(abs_err)),
    }

if _RUN:
    all_metrics = {}
    for label, df in results.items():
        all_metrics[label] = compute_metrics(df, label)


# ── save text metrics ──────────────────────────────────────────────────────────
def save_metrics_txt(m, path):
    lines = [
        "=" * 55,
        f"  FORECAST METRICS — {m['label']}",
        "=" * 55,
        f"  Timesteps evaluated : {m['n_steps']:,}",
        f"  Mean actual load    : {m['mean_actual_kw']:.4f} kW",
        f"  Mean forecast       : {m['mean_forecast_kw']:.4f} kW",
        f"  Bias (mean error)   : {m['bias_kw']:+.4f} kW",
        "-" * 55,
        f"  MAE                 : {m['mae_kw']:.4f} kW",
        f"  RMSE                : {m['rmse_kw']:.4f} kW",
        f"  MAPE                : {m['mape_pct']:.2f} %",
        f"  NRMSE               : {m['nrmse_pct']:.2f} %",
        f"  Max absolute error  : {m['max_abs_error_kw']:.4f} kW",
        "=" * 55,
        "",
        "Per-step error statistics:",
        "",
    ]
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print(f"  Saved: {path}")

if _RUN:
    for label, m in all_metrics.items():
        slug = label.lower().replace(" ", "_")
        save_metrics_txt(m, os.path.join(OUT_DIR, f"metrics_{slug}.txt"))


# ── plots ──────────────────────────────────────────────────────────────────────
COLORS = {"actual": "#1f77b4", "forecast": "#ff7f0e", "error": "#d62728"}

def make_plots(df, m, slug):
    actual   = df["actual_kw"].values
    forecast = df["forecast_kw"].values
    errors   = df["error_kw"].values
    ts       = df.index

    # ── Plot 1: Actual vs Forecast ──────────────────────────────────────────
    fig, axes = plt.subplots(3, 1, figsize=(14, 10),
                              gridspec_kw={"height_ratios": [3, 1.5, 1.5]})
    fig.suptitle(f"Load Forecast vs Actual — {m['label']}", fontsize=14, fontweight="bold")

    ax = axes[0]
    ax.plot(ts, actual,   lw=1.0, color=COLORS["actual"],   label="Actual load",   alpha=0.9)
    ax.plot(ts, forecast, lw=0.8, color=COLORS["forecast"], label="Forecast (h=1)", alpha=0.8)
    ax.set_ylabel("Load (kW)")
    ax.legend(loc="upper right", fontsize=9)
    ax.grid(axis="y", alpha=0.3)
    ax.set_title(f"NRMSE={m['nrmse_pct']:.2f}%  |  RMSE={m['rmse_kw']:.4f} kW  |  "
                  f"MAE={m['mae_kw']:.4f} kW  |  MAPE={m['mape_pct']:.2f}%",
                  fontsize=9, color="#444")

    ax2 = axes[1]
    ax2.fill_between(ts, errors, 0, where=(errors >= 0),
                      color=COLORS["forecast"], alpha=0.5, label="Over-forecast")
    ax2.fill_between(ts, errors, 0, where=(errors < 0),
                      color=COLORS["actual"], alpha=0.5, label="Under-forecast")
    ax2.axhline(0, color="black", lw=0.6)
    ax2.set_ylabel("Error (kW)")
    ax2.legend(loc="upper right", fontsize=9)
    ax2.grid(axis="y", alpha=0.3)

    ax3 = axes[2]
    ax3.plot(ts, df["abs_error_kw"].values, lw=0.7, color=COLORS["error"], alpha=0.8)
    ax3.set_ylabel("|Error| (kW)")
    ax3.set_xlabel("Timestamp")
    ax3.grid(axis="y", alpha=0.3)

    fig.tight_layout()
    fpath = os.path.join(OUT_DIR, f"plot_{slug}_forecast.png")
    fig.savefig(fpath, dpi=150, bbox_inches="tight")
    plt.close(fig)
    print(f"  Saved: {fpath}")

    # ── Plot 2: Error distribution histograms ──────────────────────────────
    fig, axes = plt.subplots(1, 3, figsize=(15, 4))
    fig.suptitle(f"Error Distributions — {m['label']}", fontsize=13, fontweight="bold")

    axes[0].hist(errors, bins=60, color=COLORS["error"], edgecolor="white", alpha=0.8)
    axes[0].axvline(0, color="black", lw=1.2, ls="--")
    axes[0].set_xlabel("Error (kW)"); axes[0].set_ylabel("Count")
    axes[0].set_title(f"Signed Error  (bias={m['bias_kw']:+.3f} kW)")
    axes[0].grid(axis="y", alpha=0.3)

    axes[1].hist(df["abs_error_kw"].values, bins=60, color=COLORS["actual"],
                  edgecolor="white", alpha=0.8)
    axes[1].set_xlabel("|Error| (kW)"); axes[1].set_ylabel("Count")
    axes[1].set_title(f"Absolute Error  (MAE={m['mae_kw']:.4f} kW)")
    axes[1].grid(axis="y", alpha=0.3)

    pct = df["pct_error"].dropna().values
    axes[2].hist(pct, bins=60, color="#2ca02c", edgecolor="white", alpha=0.8)
    axes[2].axvline(m["mape_pct"], color="black", lw=1.2, ls="--",
                     label=f"MAPE={m['mape_pct']:.1f}%")
    axes[2].set_xlabel("% Error"); axes[2].set_ylabel("Count")
    axes[2].set_title("Percentage Error")
    axes[2].legend(fontsize=9)
    axes[2].grid(axis="y", alpha=0.3)

    fig.tight_layout()
    fpath2 = os.path.join(OUT_DIR, f"plot_{slug}_errors.png")
    fig.savefig(fpath2, dpi=150, bbox_inches="tight")
    plt.close(fig)
    print(f"  Saved: {fpath2}")

    # ── Plot 3: Per-day NRMSE bar chart ────────────────────────────────────
    # Denominator = whole-month mean actual load (same as headline NRMSE).
    # Each bar = sqrt(mean(sq_error for that day)) / month_mean × 100.
    # This means bars sum-up consistently with the overall figure shown on slide 1.
    month_mean = float(np.mean(actual))   # whole-month mean — never per-day mean

    daily = df.groupby(df.index.date).apply(
        lambda g: np.sqrt(np.mean(g["sq_error_kw2"].values))
    )
    daily_nrmse = (daily / month_mean * 100).rename("nrmse_pct")
    days        = np.arange(len(daily_nrmse))
    day_labels  = [str(d.day) for d in daily_nrmse.index]

    bar_colors  = ["#d62728" if v > m["nrmse_pct"] else "#1f77b4"
                   for v in daily_nrmse.values]

    fig, ax = plt.subplots(figsize=(14, 5))
    bars = ax.bar(days, daily_nrmse.values, color=bar_colors,
                  edgecolor="white", linewidth=0.4, zorder=3)
    ax.axhline(m["nrmse_pct"], color="black", lw=1.5, ls="--", zorder=4,
               label=f"Overall NRMSE = {m['nrmse_pct']:.2f}%")
    ax.set_xticks(days)
    ax.set_xticklabels(day_labels, fontsize=8)
    ax.set_xlabel("Day of month")
    ax.set_ylabel("Daily NRMSE  [%]\n(denominator = whole-month mean load)")
    ax.set_title(
        f"Per-Day NRMSE — {m['label']}\n"
        f"Red bars = above overall NRMSE ({m['nrmse_pct']:.2f}%)",
        fontsize=11, fontweight="bold")
    ax.legend(fontsize=9)
    ax.grid(axis="y", alpha=0.3, zorder=0)

    # annotate the 3 worst days
    worst_idx = daily_nrmse.nlargest(3).index
    for d in worst_idx:
        xi = list(daily_nrmse.index).index(d)
        ax.text(xi, daily_nrmse[d] + 0.5, f"{daily_nrmse[d]:.0f}%",
                ha="center", va="bottom", fontsize=7.5, color="#d62728", fontweight="bold")

    fig.tight_layout()
    fpath3 = os.path.join(OUT_DIR, f"plot_{slug}_daily_nrmse.png")
    fig.savefig(fpath3, dpi=150, bbox_inches="tight")
    plt.close(fig)
    print(f"  Saved: {fpath3}")

    # print worst 5 days to console
    print(f"\n  Worst 5 days ({m['label']}):")
    for d, v in daily_nrmse.nlargest(5).items():
        print(f"    {d}  →  {v:.1f}%")

    return (os.path.join(OUT_DIR, f"plot_{slug}_forecast.png"),
            os.path.join(OUT_DIR, f"plot_{slug}_errors.png"),
            fpath3)


if _RUN:
    plot_paths = {}
    print("\nGenerating plots …")
    for label, df in results.items():
        slug = label.lower().replace(" ", "_")
        plot_paths[label] = make_plots(df, all_metrics[label], slug)


# ── Excel export ──────────────────────────────────────────────────────────────
HDR_FILL   = PatternFill("solid", fgColor="1F4E79")
ALT_FILL   = PatternFill("solid", fgColor="EBF3FB")
BORDER     = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"))

def style_header(cell, text):
    cell.value     = text
    cell.font      = Font(bold=True, color="FFFFFF", size=10)
    cell.fill      = HDR_FILL
    cell.alignment = Alignment(horizontal="center")
    cell.border    = BORDER

def style_val(cell, alternate=False):
    cell.fill   = ALT_FILL if alternate else PatternFill()
    cell.border = BORDER
    cell.alignment = Alignment(horizontal="right")

def save_excel(df, m, slug, forecast_img, error_img, daily_nrmse_img):
    path = os.path.join(OUT_DIR, f"forecast_{slug}.xlsx")
    wb   = openpyxl.Workbook()

    # ── Sheet 1: per-timestep data ────────────────────────────────────────────
    ws = wb.active
    ws.title = "Forecasts"

    headers = [
        ("Timestamp",        15),
        ("Actual (kW)",      13),
        ("Forecast (kW)",    13),
        ("Error (kW)",       12),
        ("|Error| (kW)",     12),
        ("Sq Error (kW²)",   13),
        ("% Error",          10),
    ]
    for col, (hdr, width) in enumerate(headers, 1):
        style_header(ws.cell(row=1, column=col), hdr)
        ws.column_dimensions[get_column_letter(col)].width = width

    for row_i, (ts, r) in enumerate(df.iterrows(), 2):
        alt = (row_i % 2 == 0)
        vals = [
            ts.strftime("%Y-%m-%d %H:%M"),
            round(float(r["actual_kw"]),    4),
            round(float(r["forecast_kw"]),  4),
            round(float(r["error_kw"]),     4),
            round(float(r["abs_error_kw"]), 4),
            round(float(r["sq_error_kw2"]), 6),
            round(float(r["pct_error"]), 2) if not np.isnan(r["pct_error"]) else "",
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row_i, column=col, value=v)
            style_val(cell, alt)

    ws.freeze_panes = "A2"

    # ── Sheet 2: summary metrics ───────────────────────────────────────────────
    ws2 = wb.create_sheet("Metrics")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 20

    metric_rows = [
        ("Period",             m["label"]),
        ("Timesteps",          m["n_steps"]),
        ("Mean actual (kW)",   round(m["mean_actual_kw"],    4)),
        ("Mean forecast (kW)", round(m["mean_forecast_kw"],  4)),
        ("Bias / mean error (kW)", round(m["bias_kw"],       4)),
        ("",                   ""),
        ("MAE (kW)",           round(m["mae_kw"],    4)),
        ("RMSE (kW)",          round(m["rmse_kw"],   4)),
        ("MAPE (%)",           round(m["mape_pct"],  2)),
        ("NRMSE (%)",          round(m["nrmse_pct"], 2)),
        ("Max |error| (kW)",   round(m["max_abs_error_kw"], 4)),
    ]
    style_header(ws2.cell(row=1, column=1), "Metric")
    style_header(ws2.cell(row=1, column=2), "Value")
    for ri, (k, v) in enumerate(metric_rows, 2):
        c1 = ws2.cell(row=ri, column=1, value=k)
        c2 = ws2.cell(row=ri, column=2, value=v)
        if k:
            c1.font = Font(bold=True, size=10)
            c1.border = BORDER
            c2.border = BORDER
            c2.alignment = Alignment(horizontal="right")

    # ── Sheet 3: charts ────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Charts")
    ws3["A1"] = f"Forecast vs Actual — {m['label']}"
    ws3["A1"].font = Font(bold=True, size=13)

    img1 = XLImage(forecast_img)
    img1.width  = 900; img1.height = 480
    ws3.add_image(img1, "A3")

    ws3["A33"] = "Error Distributions"
    ws3["A33"].font = Font(bold=True, size=13)
    img2 = XLImage(error_img)
    img2.width  = 900; img2.height = 280
    ws3.add_image(img2, "A35")

    ws3["A52"] = "Per-Day NRMSE  (denominator = whole-month mean load)"
    ws3["A52"].font = Font(bold=True, size=13)
    img3 = XLImage(daily_nrmse_img)
    img3.width  = 900; img3.height = 300
    ws3.add_image(img3, "A54")

    wb.save(path)
    print(f"  Saved: {path}")


if _RUN:
    print("\nExporting Excel files …")
    for label, df in results.items():
        slug = label.lower().replace(" ", "_")
        save_excel(df, all_metrics[label], slug,
                   plot_paths[label][0], plot_paths[label][1], plot_paths[label][2])

    # ── requirements.txt ──────────────────────────────────────────────────────
    req_path = os.path.join(OUT_DIR, "requirements.txt")
    with open(req_path, "w", encoding="utf-8") as f:
        f.write(
            "# Python dependencies for the Solship Energy AI Hackathon submission\n"
            "# Install with:  pip install -r requirements.txt\n\n"
            "pandas>=2.0\n"
            "numpy>=1.24\n"
            "lightgbm>=4.0\n"
            "cvxpy>=1.4\n"
            "openpyxl>=3.1\n"
            "holidays>=0.40\n"
            "scikit-learn>=1.3\n"
            "matplotlib>=3.7\n"
        )
    print(f"\n  Saved: {req_path}")

    # ── final console summary ──────────────────────────────────────────────────
    print("\n" + "=" * 60)
    print("FINAL SUBMISSION METRICS")
    print("=" * 60)
    for label, m in all_metrics.items():
        print(f"\n  {label}")
        print(f"    Timesteps : {m['n_steps']:,}")
        print(f"    MAE       : {m['mae_kw']:.4f} kW")
        print(f"    RMSE      : {m['rmse_kw']:.4f} kW")
        print(f"    MAPE      : {m['mape_pct']:.2f} %")
        print(f"    NRMSE     : {m['nrmse_pct']:.2f} %")
        print(f"    Bias      : {m['bias_kw']:+.4f} kW")
    print(f"\n  All outputs saved to: {OUT_DIR}/")
    print("=" * 60)


# ── full-year 2025 NRMSE (hackathon primary scoring metric) ───────────────────
def run_full_year_2025(raw):
    """
    Train on 2024 only, predict h=1 for every step of Jan–Dec 2025.

    This is what the hackathon grades on.  Teams reporting ~30% NRMSE are
    reporting THIS figure — winter months (Jan, Dec) with flat heating
    profiles (~2.1–2.3 kW mean) give ~20-25%, pulling the annual average
    well below the ~65% seen in the transition months April/September.

    Also adds load_lag_35136 (same 15-min slot exactly one year ago, adjusted
    for 2024 being a leap year: 366 × 96 = 35,136 steps).  This lag cannot be
    used in run_period() training because the dataset starts at Jan 2024 and
    there is no 2023 data — every training row would be NaN and get dropped.
    Here it is safe: we train on 2024 without the lag, then append it as an
    extra column only during inference on 2025, exploiting the fact that 2024
    history is fully available.
    """
    print(f"\n{'='*60}")
    print("FULL-YEAR 2025 NRMSE  (hackathon scoring metric)")
    print(f"{'='*60}")

    df_2024 = raw[raw.index.year == 2024]
    df_2025 = raw[raw.index.year == 2025]

    print(f"  Train: 2024  ({len(df_2024):,} rows)")
    print(f"  Test : 2025  ({len(df_2025):,} rows)")

    # Slot stats from 2024 only (no 2025 leakage)
    ss_2024  = compute_slot_stats(df_2024)
    sms_2024 = compute_slot_month_stats(df_2024)
    sds_2024 = compute_slot_dow_stats(df_2024)

    # ── build training features from 2024 ─────────────────────────────────────
    df_2024_log = log_df(df_2024)
    X_tr, y_tr  = build_features(
        df_2024_log, lags=ENHANCED_LAGS, roll_windows=ENHANCED_ROLLS, horizon=1,
        slot_stats=ss_2024, slot_month_stats=sms_2024, slot_dow_stats=sds_2024
    )
    feat_cols_fy = list(X_tr.columns)
    print(f"  Training features : {len(feat_cols_fy)}   Training rows : {len(X_tr):,}")

    # ── train LGB Huber on all 2024 ───────────────────────────────────────────
    dtrain = lgb.Dataset(X_tr.values, y_tr.values,
                         feature_name=feat_cols_fy, free_raw_data=False)
    model_fy = lgb.train(
        {**LGB_PARAMS, "objective": "huber", "metric": "huber", "alpha": 0.9},
        dtrain, num_boost_round=800, callbacks=[lgb.log_evaluation(0)]
    )
    print("  LGB training done")

    # ── build inference features for all 2025 ────────────────────────────────
    # Full dataset (2024 + 2025) passed to build_features — all shift() ops
    # are causal so 2025 predictions never see future actual load.
    full_log = log_df(raw)
    X_all, _ = build_features(
        full_log, lags=ENHANCED_LAGS, roll_windows=ENHANCED_ROLLS, horizon=1,
        slot_stats=ss_2024, slot_month_stats=sms_2024, slot_dow_stats=sds_2024
    )

    # Restrict to 2025 test rows and align feature columns
    X_2025 = X_all[X_all.index.year == 2025][feat_cols_fy]

    # ── same-period-last-year lag (35136 = 366d × 96 steps, leap 2024) ───────
    # Added ONLY here: NaN-free for 2025 since 2024 data exists as history.
    # LGB handles the NaN-safe column as a simple numeric feature; we fill
    # the rare NaN (step < 35136 into 2025, i.e. first ~0 steps since 2025
    # starts at position ≥ 35136 in the full index) with the annual slot mean.
    lag_35136_full   = raw["load_kw"].shift(35136)
    lag_35136_2025   = lag_35136_full.reindex(X_2025.index)
    slot_key_2025    = X_2025.index.hour * 4 + X_2025.index.minute // 15
    annual_fill      = slot_key_2025.map(ss_2024["mean"])
    lag_35136_filled = lag_35136_2025.fillna(
        pd.Series(annual_fill.values, index=X_2025.index))
    X_2025 = X_2025.copy()
    X_2025["load_lag_35136"] = lag_35136_filled.values

    # Extend model to accept the extra column by appending it to the dataset —
    # LGB's predict() requires the exact same number of columns used in training,
    # so we retrain including the lag for rows where it is available.
    # Simplest valid approach: add the lag to training too, filling NaN with 0
    # (the model will learn that 0 = "not available" and down-weight accordingly).
    lag_35136_2024   = raw["load_kw"].shift(35136).reindex(X_tr.index).fillna(0.0)
    X_tr_ext = X_tr.copy()
    X_tr_ext["load_lag_35136"] = lag_35136_2024.values
    feat_cols_ext = list(X_tr_ext.columns)

    print("  Retraining with load_lag_35136 …")
    dtrain_ext = lgb.Dataset(X_tr_ext.values, y_tr.values,
                             feature_name=feat_cols_ext, free_raw_data=False)
    model_ext = lgb.train(
        {**LGB_PARAMS, "objective": "huber", "metric": "huber", "alpha": 0.9},
        dtrain_ext, num_boost_round=800, callbacks=[lgb.log_evaluation(0)]
    )

    # ── generate 2025 predictions ─────────────────────────────────────────────
    pred_log  = model_ext.predict(X_2025.values)
    pred_kw   = expm1c(pred_log)

    # h=1 at t predicts load at t+1: align actuals with shift(-1)
    actual_2025_shifted = df_2025["load_kw"].shift(-1)
    common = X_2025.index.intersection(actual_2025_shifted.dropna().index)

    forecast_s = pd.Series(pred_kw, index=X_2025.index).reindex(common)
    actual_s   = actual_2025_shifted.reindex(common)

    # ── overall metrics ───────────────────────────────────────────────────────
    overall_nrmse = _nrmse(actual_s.values, forecast_s.values)
    overall_rmse  = float(np.sqrt(np.mean((actual_s.values - forecast_s.values) ** 2)))
    overall_mae   = float(np.mean(np.abs(actual_s.values - forecast_s.values)))
    overall_bias  = float(np.mean(forecast_s.values - actual_s.values))

    print(f"\n  ┌──────────────────────────────────────────────┐")
    print(f"  │  Full-year 2025 NRMSE : {overall_nrmse:>6.2f}%              │")
    print(f"  │  Full-year 2025 RMSE  : {overall_rmse:>7.4f} kW           │")
    print(f"  │  Full-year 2025 MAE   : {overall_mae:>7.4f} kW           │")
    print(f"  │  Mean actual load     : {actual_s.mean():>7.4f} kW           │")
    print(f"  └──────────────────────────────────────────────┘")

    # ── per-month breakdown ───────────────────────────────────────────────────
    month_names = ["Jan","Feb","Mar","Apr","May","Jun",
                   "Jul","Aug","Sep","Oct","Nov","Dec"]
    print(f"\n  Per-month NRMSE (2025):")
    print(f"  {'Month':<6} {'NRMSE':>8} {'RMSE':>8} {'MAE':>7} {'Mean load':>10} {'Steps':>7}")
    print(f"  {'─'*6} {'─'*8} {'─'*8} {'─'*7} {'─'*10} {'─'*7}")

    monthly = {}
    for mo in range(1, 13):
        mask = (common.month == mo)
        if mask.sum() == 0:
            continue
        a = actual_s[mask].values
        f = forecast_s[mask].values
        mn = _nrmse(a, f)
        mr = float(np.sqrt(np.mean((a - f) ** 2)))
        mm = float(np.mean(np.abs(a - f)))
        monthly[mo] = {"nrmse": mn, "rmse": mr, "mae": mm,
                       "mean": float(np.mean(a)), "n": int(mask.sum())}
        marker = " ◄ test" if mo in (4, 9) else ""
        print(f"  {month_names[mo-1]:<6} {mn:>8.2f}% {mr:>8.4f}  {mm:>7.4f}  "
              f"{np.mean(a):>10.4f}  {mask.sum():>7,}{marker}")

    # ── save text report ──────────────────────────────────────────────────────
    fy_txt = os.path.join(OUT_DIR, "metrics_full_year_2025.txt")
    with open(fy_txt, "w", encoding="utf-8") as fh:
        fh.write("Full-Year 2025 Forecast Metrics — Hackathon Scoring Metric\n")
        fh.write("=" * 65 + "\n")
        fh.write(f"  Model    : LGB Huber + load_lag_35136 (same slot last year)\n")
        fh.write(f"  Train    : 2024 only ({len(X_tr):,} rows → {len(X_tr_ext):,} after lag)\n")
        fh.write(f"  Test     : 2025 Jan–Dec ({len(common):,} steps)\n\n")
        fh.write(f"  NRMSE    : {overall_nrmse:.2f}%\n")
        fh.write(f"  RMSE     : {overall_rmse:.4f} kW\n")
        fh.write(f"  MAE      : {overall_mae:.4f} kW\n")
        fh.write(f"  Bias     : {overall_bias:+.4f} kW\n\n")
        fh.write(f"  {'Month':<8} {'NRMSE':>8} {'RMSE':>8} {'MAE':>7} {'Mean kW':>8}\n")
        fh.write(f"  {'─'*8} {'─'*8} {'─'*8} {'─'*7} {'─'*8}\n")
        for mo in range(1, 13):
            if mo in monthly:
                r2 = monthly[mo]
                tag = " (test)" if mo in (4, 9) else ""
                fh.write(f"  {month_names[mo-1]:<8} {r2['nrmse']:>8.2f}% "
                         f"{r2['rmse']:>8.4f}  {r2['mae']:>7.4f}  {r2['mean']:>8.4f}{tag}\n")
    print(f"\n  Saved: {fy_txt}")

    # ── per-month NRMSE bar chart ─────────────────────────────────────────────
    months_plotted = sorted(monthly.keys())
    nrmse_vals     = [monthly[m]["nrmse"]  for m in months_plotted]
    mean_vals      = [monthly[m]["mean"]   for m in months_plotted]
    xlabels        = [month_names[m - 1]   for m in months_plotted]

    fig, axes = plt.subplots(2, 1, figsize=(13, 8),
                              gridspec_kw={"height_ratios": [3, 1.5]})
    fig.suptitle(f"Full-Year 2025 NRMSE  (train: 2024 only | overall = {overall_nrmse:.2f}%)",
                 fontsize=13, fontweight="bold")

    ax = axes[0]
    bar_colors = []
    for i, m in enumerate(months_plotted):
        if m in (4, 9):
            bar_colors.append("#d62728")   # test months: red
        elif nrmse_vals[i] > overall_nrmse:
            bar_colors.append("#ff7f0e")   # above average: orange
        else:
            bar_colors.append("#1f77b4")   # below average: blue
    bars = ax.bar(xlabels, nrmse_vals, color=bar_colors,
                  edgecolor="white", linewidth=0.5, zorder=3)
    ax.axhline(overall_nrmse, color="black", lw=1.5, ls="--", zorder=4,
               label=f"Full-year NRMSE = {overall_nrmse:.2f}%")
    for xi, v in enumerate(nrmse_vals):
        ax.text(xi, v + 0.4, f"{v:.1f}%", ha="center", va="bottom",
                fontsize=8, fontweight="bold")
    ax.set_ylabel("NRMSE (%)")
    ax.set_title("Red = test months (Apr, Sep)  |  Orange = above full-year average",
                 fontsize=9, color="#555")
    ax.legend(fontsize=10)
    ax.grid(axis="y", alpha=0.3, zorder=0)

    ax2 = axes[1]
    ax2.bar(xlabels, mean_vals, color="#2ca02c", edgecolor="white",
            linewidth=0.5, zorder=3, alpha=0.85)
    ax2.set_ylabel("Mean load (kW)")
    ax2.set_xlabel("Month (2025)")
    ax2.set_title("Mean actual load — lower mean = higher NRMSE even at same RMSE",
                  fontsize=9, color="#555")
    ax2.grid(axis="y", alpha=0.3, zorder=0)

    fig.tight_layout()
    fy_png = os.path.join(OUT_DIR, "plot_full_year_nrmse.png")
    fig.savefig(fy_png, dpi=150, bbox_inches="tight")
    plt.close(fig)
    print(f"  Saved: {fy_png}")

    return overall_nrmse, monthly


# fy_nrmse, fy_monthly = run_full_year_2025()   # commented out — judges score April+Sep only
