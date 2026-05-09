"""
Final submission: h=1 load forecast for April 2025 and September 2025.

Each period is predicted using ALL data that came before it:
  April     → train Jan 2024 – Jan 2025 | val Feb–Mar 2025
  September → train Jan 2024 – Jun 2025 | val Jul–Aug 2025

Outputs saved to  submission_forecast/
  requirements.txt
  metrics_april_2025.txt
  metrics_september_2025.txt
  forecast_april_2025.xlsx    (data + per-step errors + summary + embedded charts)
  forecast_september_2025.xlsx
  plot_april_forecast.png
  plot_april_errors.png
  plot_september_forecast.png
  plot_september_errors.png
"""

import sys, os, warnings, time
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "solship"))
warnings.filterwarnings("ignore")

import numpy as np
import pandas as pd
import lightgbm as lgb
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import holidays as hol_pkg
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from features import build_features, compute_slot_stats

# ── constants ──────────────────────────────────────────────────────────────────
OUT_DIR          = "submission_forecast"
F1, F2, F3       = 0.2540, 0.2682, 0.2440
ITALIAN_HOLIDAYS = hol_pkg.Italy(years=[2024, 2025])
# Extended lags: add 2-week (1344), 3-week (2016), 4-week (2688) for transition months
ENHANCED_LAGS    = [0, 1, 2, 3, 4, 5, 6, 7, 8, 48, 96, 192, 336, 672, 1344, 2016, 2688]
ENHANCED_ROLLS   = [4, 8, 16, 96, 672, 1344]   # added 2-week rolling
SPARSE_HORIZONS  = [1, 2, 4, 8, 16, 24, 48, 72, 96, 144, 192]
H_MAX            = 192

LGB_PARAMS = {
    "verbose": -1, "num_threads": -1,
    "learning_rate": 0.0272, "num_leaves": 167,
    "min_data_in_leaf": 135, "feature_fraction": 0.862,
    "bagging_fraction": 0.656, "bagging_freq": 5,
    "lambda_l2": 0.078, "lambda_l1": 0.150,
}

log1p  = np.log1p
expm1c = lambda x: np.expm1(x).clip(min=0)

os.makedirs(OUT_DIR, exist_ok=True)

# ── helpers ────────────────────────────────────────────────────────────────────
def compute_buy_price(ts):
    dow = ts.dt.dayofweek; h = ts.dt.hour
    hol = ts.dt.date.map(lambda d: d in ITALIAN_HOLIDAYS)
    f3  = (dow == 6) | hol | (~(dow == 6) & ~hol & ((h < 7) | (h >= 23)))
    f2  = (~f3) & (((dow == 5) & (h >= 7) & (h < 23)) |
                   (~(dow == 5) & ~(dow == 6) & ~hol & ((h == 7) | ((h >= 19) & (h < 23)))))
    p   = pd.Series(F1, index=ts.index, dtype=float); p[f2] = F2; p[f3] = F3
    return p

def log_df(df):
    d = df.copy(); d["load_kw"] = log1p(d["load_kw"]); return d

def make_Xy(df_buf, df_target, horizon, feat_cols=None, slot_stats=None):
    X, y = build_features(df_buf, lags=ENHANCED_LAGS,
                           roll_windows=ENHANCED_ROLLS, horizon=horizon,
                           slot_stats=slot_stats)
    idx  = df_target.index.intersection(X.index)
    X, y = X.loc[idx], y.loc[idx]
    if feat_cols is not None:
        X = X[feat_cols]
    return X, y

def safe_mape(actual, forecast):
    mask = actual > 0.01
    if mask.sum() == 0:
        return float("nan")
    return float(np.mean(np.abs((actual[mask] - forecast[mask]) / actual[mask])) * 100)

# ── load & clean data ──────────────────────────────────────────────────────────
print("=" * 60)
print("LOADING DATA")
print("=" * 60)

raw = pd.read_csv("data/ENERGY_Hackathon_DataSet(Sheet1).csv", sep=";", decimal=",")
raw["timestamp"] = pd.to_datetime(raw["timestamp"])
raw = (raw.sort_values("timestamp")
          .drop_duplicates(subset="timestamp", keep="first")
          .reset_index(drop=True))
raw = raw.rename(columns={"load_p": "load_kw", "pv_p": "pv_kw",
                           "Selling_price_eur_kwh": "sell_price",
                           "battery_p": "p_battery_kw", "grid_p": "grid_kw"})
raw["sell_price"] = raw["sell_price"].ffill()
raw["buy_price"]  = compute_buy_price(raw["timestamp"])
idx = pd.date_range(raw["timestamp"].min(), raw["timestamp"].max(), freq="15min")
raw = raw.set_index("timestamp").reindex(idx).ffill()
raw["load_kw"] = raw["load_kw"].clip(lower=0)
raw["pv_kw"]   = raw["pv_kw"].clip(lower=0)
print(f"  Rows loaded: {len(raw):,}  ({raw.index[0].date()} → {raw.index[-1].date()})")


# ── training / prediction for one period ──────────────────────────────────────
def _nrmse(actual, forecast):
    rmse = np.sqrt(np.mean((actual - forecast) ** 2))
    return rmse / np.mean(actual) * 100 if np.mean(actual) > 1e-9 else float("nan")


def run_period(month_num, label, month_start, month_end):
    """
    Train on ALL data before the target month.

    ES validation uses the 8 weeks immediately before the test month — this
    matches the seasonal profile of the test month (spring for April, late
    summer for September) and gives much better early-stopping signal than
    a fixed winter holdout.

    Blend weight alpha (Huber fraction) is optimized on the ES val set.

    Returns DataFrame with per-timestep actuals, forecasts, and errors.
    """
    print(f"\n{'='*60}")
    print(f"PERIOD: {label}  ({month_start.date()} – {month_end.date()})")
    print(f"{'='*60}")
    t0 = time.time()

    train_end = month_start - pd.Timedelta("15min")
    df_train  = raw[raw.index <= train_end]
    df_test   = raw[(raw.index >= month_start) & (raw.index <= month_end)]

    # ES validation: 8 weeks closest to the test month (best seasonal proxy)
    # This is training data only — test month is never seen
    es_split  = train_end - pd.Timedelta(days=56)
    df_tr_es  = df_train[df_train.index <= es_split]
    df_val_es = df_train[df_train.index >  es_split]

    print(f"  Train : {df_train.index[0].date()} – {df_train.index[-1].date()}  "
          f"({len(df_train):,} rows)")
    print(f"  ES val: {df_val_es.index[0].date()} – {df_val_es.index[-1].date()}  "
          f"({len(df_val_es):,} rows, early-stop + blend tuning)")
    print(f"  Test  : {df_test.index[0].date()} – {df_test.index[-1].date()}  "
          f"({len(df_test):,} rows)")

    # ── compute per-slot stats from training data only (prevents test leakage) ──
    train_slot_stats = compute_slot_stats(df_train)
    tr_es_slot_stats = compute_slot_stats(df_tr_es)

    # ── log-transform datasets ──
    df_tr_log    = log_df(df_train)
    df_tr_es_log = log_df(df_tr_es)
    # buffer for ES val: include tail of df_tr_es so lags can be computed
    buf_tail     = max(ENHANCED_LAGS)
    df_va_buf    = log_df(pd.concat([df_tr_es.iloc[-buf_tail:], df_val_es]))

    # Build feature columns from full training set with training slot stats
    X_tr1, _ = make_Xy(df_tr_log, df_train, 1, slot_stats=train_slot_stats)
    feat_cols = list(X_tr1.columns)

    models_hub = {}; models_qnt = {}

    print(f"  Training {len(SPARSE_HORIZONS)} horizons × 2 models …")
    t_tr = time.time()
    for h in SPARSE_HORIZONS:
        # Early-stop split uses tr_es slot stats to avoid any val leakage
        X_es, y_es = make_Xy(df_tr_es_log, df_tr_es, h, feat_cols, slot_stats=tr_es_slot_stats)
        X_va, y_va = make_Xy(df_va_buf,    df_val_es, h, feat_cols, slot_stats=tr_es_slot_stats)
        d_es  = lgb.Dataset(X_es.values, y_es.values, feature_name=feat_cols, free_raw_data=False)
        d_va  = lgb.Dataset(X_va.values, y_va.values, feature_name=feat_cols,
                             reference=d_es, free_raw_data=False)
        cbs_es = [lgb.early_stopping(80, verbose=False), lgb.log_evaluation(0)]
        m_es = lgb.train(
            {**LGB_PARAMS, "objective": "huber", "metric": "huber", "alpha": 0.9},
            d_es, num_boost_round=2000, valid_sets=[d_va], callbacks=cbs_es)
        best_rounds = max(m_es.best_iteration, 50)

        # Final model: all training data, fixed rounds
        X_tr, y_tr = make_Xy(df_tr_log, df_train, h, feat_cols, slot_stats=train_slot_stats)
        dtrain = lgb.Dataset(X_tr.values, y_tr.values, feature_name=feat_cols, free_raw_data=False)
        cbs = [lgb.log_evaluation(0)]
        models_hub[h] = lgb.train(
            {**LGB_PARAMS, "objective": "huber", "metric": "huber", "alpha": 0.9},
            dtrain, num_boost_round=best_rounds, callbacks=cbs)
        models_qnt[h] = lgb.train(
            {**LGB_PARAMS, "objective": "quantile", "metric": "quantile", "alpha": 0.50},
            dtrain, num_boost_round=best_rounds, callbacks=cbs)
    print(f"  Training done in {time.time()-t_tr:.1f}s")

    # ── find optimal blend weight on ES val set (h=1 only) ────────────────────
    X_va1, y_va1 = make_Xy(df_va_buf, df_val_es, 1, feat_cols, slot_stats=tr_es_slot_stats)
    val_hub1 = expm1c(models_hub[1].predict(X_va1.values))
    val_qnt1 = expm1c(models_qnt[1].predict(X_va1.values))
    y_va1_act = np.expm1(y_va1.values).clip(min=0)   # y_va1 was log-transformed target

    best_alpha, best_nrmse = 0.5, float("inf")
    for alpha in np.linspace(0, 1, 21):
        blend_val = alpha * val_hub1 + (1 - alpha) * val_qnt1
        n = _nrmse(y_va1_act, blend_val)
        if n < best_nrmse:
            best_nrmse, best_alpha = n, alpha
    print(f"  Optimal blend alpha (Huber fraction): {best_alpha:.2f}  "
          f"(val NRMSE={best_nrmse:.2f}%)")

    # ── generate forecasts for the test month using training slot stats ────────
    buf_data  = raw[raw.index < month_start]
    pad_idx   = pd.date_range(df_test.index[-1] + pd.Timedelta("15min"),
                               periods=H_MAX, freq="15min")
    pad_df    = pd.concat([df_test.iloc[[-1]]] * H_MAX)
    pad_df.index = pad_idx
    df_te_buf = log_df(pd.concat([buf_data, df_test, pad_df]))

    pred_hub = {}; pred_qnt = {}
    for h in SPARSE_HORIZONS:
        # Pass training slot_stats — test data in df_te_buf does NOT influence features
        X, _ = make_Xy(df_te_buf, df_test, h, feat_cols, slot_stats=train_slot_stats)
        pred_hub[h] = pd.Series(expm1c(models_hub[h].predict(X.values)), index=X.index)
        pred_qnt[h] = pd.Series(expm1c(models_qnt[h].predict(X.values)), index=X.index)

    common = pred_hub[SPARSE_HORIZONS[0]].index
    for h in SPARSE_HORIZONS[1:]:
        common = common.intersection(pred_hub[h].index)
    common = common[(common >= month_start) & (common <= month_end)]

    sparse_arr = np.array(SPARSE_HORIZONS)
    hub_mat = np.column_stack([pred_hub[h].loc[common].values for h in SPARSE_HORIZONS])
    qnt_mat = np.column_stack([pred_qnt[h].loc[common].values for h in SPARSE_HORIZONS])
    blend   = best_alpha * hub_mat + (1 - best_alpha) * qnt_mat

    # h=1 column only (h=1 is directly in SPARSE_HORIZONS, no interpolation needed)
    h1_blend = np.array([np.interp(1, sparse_arr, blend[i])
                          for i in range(len(common))]).clip(min=0)

    forecast_h1 = pd.Series(h1_blend, index=common, name="forecast_kw")

    # h=1 at t predicts load at t+1: align actuals with correct shift(-1)
    actual_h1 = df_test["load_kw"].shift(-1).reindex(common).dropna()
    forecast_h1 = forecast_h1.reindex(actual_h1.index)

    # ── per-timestep error table ──────────────────────────────────────────────
    result = pd.DataFrame({
        "actual_kw":    actual_h1.values,
        "forecast_kw":  forecast_h1.values,
    }, index=actual_h1.index)

    result["error_kw"]     = result["forecast_kw"] - result["actual_kw"]
    result["abs_error_kw"] = result["error_kw"].abs()
    result["sq_error_kw2"] = result["error_kw"] ** 2
    mask_nonzero           = result["actual_kw"] > 0.01
    result["pct_error"]    = np.nan
    result.loc[mask_nonzero, "pct_error"] = (
        result.loc[mask_nonzero, "abs_error_kw"] /
        result.loc[mask_nonzero, "actual_kw"] * 100
    )

    print(f"  Forecast timesteps: {len(result):,}")
    print(f"  Total elapsed: {time.time()-t0:.1f}s")

    # ── feature importance (h=1 Huber model, gain-based) ──────────────────────
    fi = pd.Series(
        models_hub[1].feature_importance(importance_type="gain"),
        index=feat_cols
    ).sort_values(ascending=False)

    top_n = 30
    slug  = label.lower().replace(" ", "_")

    # text file: full ranked list
    fi_txt = os.path.join(OUT_DIR, f"feature_importance_{slug}.txt")
    with open(fi_txt, "w", encoding="utf-8") as f:
        f.write(f"Feature importance (gain) — {label}\n")
        f.write("=" * 50 + "\n")
        for rank, (feat, val) in enumerate(fi.items(), 1):
            f.write(f"  {rank:3d}.  {feat:<35s}  {val:,.0f}\n")
    print(f"  Saved: {fi_txt}")

    # bar chart: top N
    fig, ax = plt.subplots(figsize=(9, top_n * 0.28 + 1.5))
    top = fi.head(top_n)
    ax.barh(top.index[::-1], top.values[::-1], color="#1f77b4", edgecolor="white")
    ax.set_xlabel("Gain (importance)")
    ax.set_title(f"Top {top_n} features by gain — {label}", fontsize=11, fontweight="bold")
    ax.grid(axis="x", alpha=0.3)
    fig.tight_layout()
    fi_png = os.path.join(OUT_DIR, f"feature_importance_{slug}.png")
    fig.savefig(fi_png, dpi=150, bbox_inches="tight")
    plt.close(fig)
    print(f"  Saved: {fi_png}")

    print(f"\n  Top 10 features ({label}):")
    for feat, val in fi.head(10).items():
        print(f"    {feat:<35s}  {val:,.0f}")

    return result


# ── run both periods ───────────────────────────────────────────────────────────
results = {}
for m_num, label, m_start, m_end in [
    (4,  "April 2025",     pd.Timestamp("2025-04-01"), pd.Timestamp("2025-04-30 23:45")),
    (9,  "September 2025", pd.Timestamp("2025-09-01"), pd.Timestamp("2025-09-30 23:45")),
]:
    results[label] = run_period(m_num, label, m_start, m_end)



# ── compute summary metrics ────────────────────────────────────────────────────
def compute_metrics(df, label):
    actual   = df["actual_kw"].values
    forecast = df["forecast_kw"].values
    errors   = df["error_kw"].values
    abs_err  = df["abs_error_kw"].values
    sq_err   = df["sq_error_kw2"].values

    mae_v   = float(np.mean(abs_err))
    rmse_v  = float(np.sqrt(np.mean(sq_err)))
    mape_v  = float(df["pct_error"].mean())
    nrmse_v = rmse_v / float(np.mean(actual)) * 100 if np.mean(actual) > 1e-9 else float("nan")

    return {
        "label":     label,
        "n_steps":   len(df),
        "mae_kw":    mae_v,
        "rmse_kw":   rmse_v,
        "mape_pct":  mape_v,
        "nrmse_pct": nrmse_v,
        "mean_actual_kw": float(np.mean(actual)),
        "mean_forecast_kw": float(np.mean(forecast)),
        "bias_kw":   float(np.mean(errors)),
        "max_abs_error_kw": float(np.max(abs_err)),
    }

all_metrics = {}
for label, df in results.items():
    all_metrics[label] = compute_metrics(df, label)


# ── save text metrics ──────────────────────────────────────────────────────────
def save_metrics_txt(m, path):
    lines = [
        "=" * 55,
        f"  FORECAST METRICS — {m['label']}",
        "=" * 55,
        f"  Timesteps evaluated : {m['n_steps']:,}",
        f"  Mean actual load    : {m['mean_actual_kw']:.4f} kW",
        f"  Mean forecast       : {m['mean_forecast_kw']:.4f} kW",
        f"  Bias (mean error)   : {m['bias_kw']:+.4f} kW",
        "-" * 55,
        f"  MAE                 : {m['mae_kw']:.4f} kW",
        f"  RMSE                : {m['rmse_kw']:.4f} kW",
        f"  MAPE                : {m['mape_pct']:.2f} %",
        f"  NRMSE               : {m['nrmse_pct']:.2f} %",
        f"  Max absolute error  : {m['max_abs_error_kw']:.4f} kW",
        "=" * 55,
        "",
        "Per-step error statistics:",
        "",
    ]
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print(f"  Saved: {path}")

for label, m in all_metrics.items():
    slug = label.lower().replace(" ", "_")
    save_metrics_txt(m, os.path.join(OUT_DIR, f"metrics_{slug}.txt"))


# ── plots ──────────────────────────────────────────────────────────────────────
COLORS = {"actual": "#1f77b4", "forecast": "#ff7f0e", "error": "#d62728"}

def make_plots(df, m, slug):
    actual   = df["actual_kw"].values
    forecast = df["forecast_kw"].values
    errors   = df["error_kw"].values
    ts       = df.index

    # ── Plot 1: Actual vs Forecast ──────────────────────────────────────────
    fig, axes = plt.subplots(3, 1, figsize=(14, 10),
                              gridspec_kw={"height_ratios": [3, 1.5, 1.5]})
    fig.suptitle(f"Load Forecast vs Actual — {m['label']}", fontsize=14, fontweight="bold")

    ax = axes[0]
    ax.plot(ts, actual,   lw=1.0, color=COLORS["actual"],   label="Actual load",   alpha=0.9)
    ax.plot(ts, forecast, lw=0.8, color=COLORS["forecast"], label="Forecast (h=1)", alpha=0.8)
    ax.set_ylabel("Load (kW)")
    ax.legend(loc="upper right", fontsize=9)
    ax.grid(axis="y", alpha=0.3)
    ax.set_title(f"NRMSE={m['nrmse_pct']:.2f}%  |  RMSE={m['rmse_kw']:.4f} kW  |  "
                  f"MAE={m['mae_kw']:.4f} kW  |  MAPE={m['mape_pct']:.2f}%",
                  fontsize=9, color="#444")

    ax2 = axes[1]
    ax2.fill_between(ts, errors, 0, where=(errors >= 0),
                      color=COLORS["forecast"], alpha=0.5, label="Over-forecast")
    ax2.fill_between(ts, errors, 0, where=(errors < 0),
                      color=COLORS["actual"], alpha=0.5, label="Under-forecast")
    ax2.axhline(0, color="black", lw=0.6)
    ax2.set_ylabel("Error (kW)")
    ax2.legend(loc="upper right", fontsize=9)
    ax2.grid(axis="y", alpha=0.3)

    ax3 = axes[2]
    ax3.plot(ts, df["abs_error_kw"].values, lw=0.7, color=COLORS["error"], alpha=0.8)
    ax3.set_ylabel("|Error| (kW)")
    ax3.set_xlabel("Timestamp")
    ax3.grid(axis="y", alpha=0.3)

    fig.tight_layout()
    fpath = os.path.join(OUT_DIR, f"plot_{slug}_forecast.png")
    fig.savefig(fpath, dpi=150, bbox_inches="tight")
    plt.close(fig)
    print(f"  Saved: {fpath}")

    # ── Plot 2: Error distribution histograms ──────────────────────────────
    fig, axes = plt.subplots(1, 3, figsize=(15, 4))
    fig.suptitle(f"Error Distributions — {m['label']}", fontsize=13, fontweight="bold")

    axes[0].hist(errors, bins=60, color=COLORS["error"], edgecolor="white", alpha=0.8)
    axes[0].axvline(0, color="black", lw=1.2, ls="--")
    axes[0].set_xlabel("Error (kW)"); axes[0].set_ylabel("Count")
    axes[0].set_title(f"Signed Error  (bias={m['bias_kw']:+.3f} kW)")
    axes[0].grid(axis="y", alpha=0.3)

    axes[1].hist(df["abs_error_kw"].values, bins=60, color=COLORS["actual"],
                  edgecolor="white", alpha=0.8)
    axes[1].set_xlabel("|Error| (kW)"); axes[1].set_ylabel("Count")
    axes[1].set_title(f"Absolute Error  (MAE={m['mae_kw']:.4f} kW)")
    axes[1].grid(axis="y", alpha=0.3)

    pct = df["pct_error"].dropna().values
    axes[2].hist(pct, bins=60, color="#2ca02c", edgecolor="white", alpha=0.8)
    axes[2].axvline(m["mape_pct"], color="black", lw=1.2, ls="--",
                     label=f"MAPE={m['mape_pct']:.1f}%")
    axes[2].set_xlabel("% Error"); axes[2].set_ylabel("Count")
    axes[2].set_title("Percentage Error")
    axes[2].legend(fontsize=9)
    axes[2].grid(axis="y", alpha=0.3)

    fig.tight_layout()
    fpath2 = os.path.join(OUT_DIR, f"plot_{slug}_errors.png")
    fig.savefig(fpath2, dpi=150, bbox_inches="tight")
    plt.close(fig)
    print(f"  Saved: {fpath2}")

    return (os.path.join(OUT_DIR, f"plot_{slug}_forecast.png"),
            os.path.join(OUT_DIR, f"plot_{slug}_errors.png"))


plot_paths = {}
print("\nGenerating plots …")
for label, df in results.items():
    slug = label.lower().replace(" ", "_")
    plot_paths[label] = make_plots(df, all_metrics[label], slug)


# ── Excel export ──────────────────────────────────────────────────────────────
HDR_FILL   = PatternFill("solid", fgColor="1F4E79")
ALT_FILL   = PatternFill("solid", fgColor="EBF3FB")
BORDER     = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"))

def style_header(cell, text):
    cell.value     = text
    cell.font      = Font(bold=True, color="FFFFFF", size=10)
    cell.fill      = HDR_FILL
    cell.alignment = Alignment(horizontal="center")
    cell.border    = BORDER

def style_val(cell, alternate=False):
    cell.fill   = ALT_FILL if alternate else PatternFill()
    cell.border = BORDER
    cell.alignment = Alignment(horizontal="right")

def save_excel(df, m, slug, forecast_img, error_img):
    path = os.path.join(OUT_DIR, f"forecast_{slug}.xlsx")
    wb   = openpyxl.Workbook()

    # ── Sheet 1: per-timestep data ────────────────────────────────────────────
    ws = wb.active
    ws.title = "Forecasts"

    headers = [
        ("Timestamp",        15),
        ("Actual (kW)",      13),
        ("Forecast (kW)",    13),
        ("Error (kW)",       12),
        ("|Error| (kW)",     12),
        ("Sq Error (kW²)",   13),
        ("% Error",          10),
    ]
    for col, (hdr, width) in enumerate(headers, 1):
        style_header(ws.cell(row=1, column=col), hdr)
        ws.column_dimensions[get_column_letter(col)].width = width

    for row_i, (ts, r) in enumerate(df.iterrows(), 2):
        alt = (row_i % 2 == 0)
        vals = [
            ts.strftime("%Y-%m-%d %H:%M"),
            round(float(r["actual_kw"]),    4),
            round(float(r["forecast_kw"]),  4),
            round(float(r["error_kw"]),     4),
            round(float(r["abs_error_kw"]), 4),
            round(float(r["sq_error_kw2"]), 6),
            round(float(r["pct_error"]), 2) if not np.isnan(r["pct_error"]) else "",
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row_i, column=col, value=v)
            style_val(cell, alt)

    ws.freeze_panes = "A2"

    # ── Sheet 2: summary metrics ───────────────────────────────────────────────
    ws2 = wb.create_sheet("Metrics")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 20

    metric_rows = [
        ("Period",             m["label"]),
        ("Timesteps",          m["n_steps"]),
        ("Mean actual (kW)",   round(m["mean_actual_kw"],    4)),
        ("Mean forecast (kW)", round(m["mean_forecast_kw"],  4)),
        ("Bias / mean error (kW)", round(m["bias_kw"],       4)),
        ("",                   ""),
        ("MAE (kW)",           round(m["mae_kw"],    4)),
        ("RMSE (kW)",          round(m["rmse_kw"],   4)),
        ("MAPE (%)",           round(m["mape_pct"],  2)),
        ("NRMSE (%)",          round(m["nrmse_pct"], 2)),
        ("Max |error| (kW)",   round(m["max_abs_error_kw"], 4)),
    ]
    style_header(ws2.cell(row=1, column=1), "Metric")
    style_header(ws2.cell(row=1, column=2), "Value")
    for ri, (k, v) in enumerate(metric_rows, 2):
        c1 = ws2.cell(row=ri, column=1, value=k)
        c2 = ws2.cell(row=ri, column=2, value=v)
        if k:
            c1.font = Font(bold=True, size=10)
            c1.border = BORDER
            c2.border = BORDER
            c2.alignment = Alignment(horizontal="right")

    # ── Sheet 3: charts ────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Charts")
    ws3["A1"] = f"Forecast vs Actual — {m['label']}"
    ws3["A1"].font = Font(bold=True, size=13)

    img1 = XLImage(forecast_img)
    img1.width  = 900; img1.height = 480
    ws3.add_image(img1, "A3")

    ws3["A33"] = "Error Distributions"
    ws3["A33"].font = Font(bold=True, size=13)
    img2 = XLImage(error_img)
    img2.width  = 900; img2.height = 280
    ws3.add_image(img2, "A35")

    wb.save(path)
    print(f"  Saved: {path}")


print("\nExporting Excel files …")
for label, df in results.items():
    slug = label.lower().replace(" ", "_")
    save_excel(df, all_metrics[label], slug,
               plot_paths[label][0], plot_paths[label][1])


# ── requirements.txt ──────────────────────────────────────────────────────────
req_path = os.path.join(OUT_DIR, "requirements.txt")
with open(req_path, "w", encoding="utf-8") as f:
    f.write(
        "# Python dependencies for the Solship Energy AI Hackathon submission\n"
        "# Install with:  pip install -r requirements.txt\n\n"
        "pandas>=2.0\n"
        "numpy>=1.24\n"
        "lightgbm>=4.0\n"
        "cvxpy>=1.4\n"
        "openpyxl>=3.1\n"
        "holidays>=0.40\n"
        "scikit-learn>=1.3\n"
        "matplotlib>=3.7\n"
    )
print(f"\n  Saved: {req_path}")


# ── final console summary ──────────────────────────────────────────────────────
print("\n" + "=" * 60)
print("FINAL SUBMISSION METRICS")
print("=" * 60)
for label, m in all_metrics.items():
    print(f"\n  {label}")
    print(f"    Timesteps : {m['n_steps']:,}")
    print(f"    MAE       : {m['mae_kw']:.4f} kW")
    print(f"    RMSE      : {m['rmse_kw']:.4f} kW")
    print(f"    MAPE      : {m['mape_pct']:.2f} %")
    print(f"    NRMSE     : {m['nrmse_pct']:.2f} %")
    print(f"    Bias      : {m['bias_kw']:+.4f} kW")

print(f"\n  All outputs saved to: {OUT_DIR}/")
print("=" * 60)
